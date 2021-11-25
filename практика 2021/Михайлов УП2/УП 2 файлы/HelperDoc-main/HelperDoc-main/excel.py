import openpyxl


class Excel:
    def __init__(self, f):
        self.book = openpyxl.open(f, read_only=True)
        self.sheet = self.book.active
        self.title_row_num = 1
        self.min_col = 0
        self.max_col = self.sheet.max_column
        self.min_row = 1
        self.max_row = self.sheet.max_row

    def set_sheet(self, num):
        """
        Установка нужного листа
        """
        is_valid = False
        if num > -1:
            self.sheet = self.book.worksheets(num)
            is_valid = True
        return is_valid

    def set_title_line_num(self, num):
        """
        Установка строки с титульником
        """
        is_valid = False
        if self.title_row_num > 0:
            self.title_row_num = num
            is_valid = True
        return is_valid

    def set_min_col(self, num):
        """
        Установка стартовой колонны
        """
        is_valid = False
        if self.min_col > -1:
            self.min_col = num
            is_valid = True
        return is_valid

    def set_max_col(self, num):
        """
        Установка конечной колонны
        """
        is_valid = False
        if self.max_col > -1:
            self.max_col = num
            is_valid = True
        return is_valid

    def get_title_row(self, title_row_num=None, min_col=None, max_col=None):
        """
        Вернет список с данными из строки с титульником
        """
        rows = list()
        if title_row_num is None:
            title_row_num = self.title_row_num
        if min_col is None:
            min_col = self.min_col
        if max_col is None:
            max_col = self.max_col

        for col in range(min_col, max_col):
            rows.append(self.sheet[title_row_num][col].value)
        return rows

    def get_row(self, row_num=None, min_col=None, max_col=None):
        """
        Вернет список с данными из строки
        """
        if row_num is None:
            row_num = self.title_row_num
        if min_col is None:
            min_col = self.min_col
        if max_col is None:
            max_col = self.max_col

        rows = list()
        for col in range(min_col, max_col):
            rows.append(self.sheet[row_num][col].value)
        return rows

    def get_rows(self, title=True, min_row=None, max_row=None, min_col=None, max_col=None):
        """
        Вернет список с данными из строк
        """
        if title is False:
            min_row = self.title_row_num+1
        if min_row is None:
            min_row = self.min_row
        if max_row is None:
            max_row = self.max_row
        if min_col is None:
            min_col = self.min_col
        if max_col is None:
            max_col = self.max_col

        rows = list()
        for row in range(min_row, max_row):
            rows.append([])
            for col in range(min_col, max_col):
                rows[-1].append(self.sheet[row][col].value)
        return rows

